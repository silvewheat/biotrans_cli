# -*- coding: utf-8 -*-
"""
Created on Wed Oct  7 10:39:50 2020

@author: YudongCai

@Email: yudongcai216@gmail.com
"""

import os
import click
CONTEXT_SETTINGS = dict(help_option_names=['-h', '--help'])
import numpy as np
import pandas as pd
from cyvcf2 import VCF
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def loadrename(rename):
    old2new = {}
    with open(rename) as f:
        for line in f:
            tline = line.strip().split()
            old2new[tline[0]] = tline[1]
    return old2new


def plot_non_trasposed():
    """
    一行一个位点，一列一个个体
    先不写了，下次加新功能的时候再写重构

    """
    pass

def plot_transposed():
    """
    一行一个个体，一列一个位点
    先不写了，下次加新功能的时候再写重构
    """
    pass


@click.command()
@click.option('--vcffile')
@click.option('--region', help='要输出的区域, 如12:1000-2000', default=None)
@click.option('--regionfile', help='要输出的多个区域，三列, chr\\tstart\\tend', default=None)
@click.option('--outlist', help='以outlist中出现最多的allel定位祖先allel(为了保持outlist中颜色尽可能一致)', default=None)
@click.option('--maf', help='只输出最小等位基因频率高于这个值的, default=0.2', type=float, default=0.2)
@click.option('--querylist', help='只输出这些个体(文本文件, 一行一个个体), 输出会按这个文件的来排序')
@click.option('--rename', help='修改输出的样本ID, 两列, [旧ID 新ID]', default=None)
@click.option('--transpose', help='行列转置，转置后一行一个个体，一列一个位点', is_flag=True, default=False)
@click.option('--outfile', help='输出文件名，须以.xlsx为后缀')
def main(vcffile, region, regionfile, outlist, maf, querylist, rename, transpose, outfile):
    """
    从vcf文件中提取指定区域存在excel表中，并根据基因型标记不同的颜色
    NOTE! 在输出结果中只保留了双等位位点
    """
    if os.path.exists(outfile):
        wb = load_workbook(outfile)
        print(f'NOTE: {outfile} already existes, the new results will added in the exist file!')
    else:
        wb = Workbook()
        del(wb['Sheet']) # 删掉初始化的空白sheet
    # 读样本ID和vcf
    querysamples = [x.strip() for x in open(querylist)]
    if outlist:
        outsamples = [x.strip() for x in open(outlist)]
        vcf_outgroup = VCF(vcffile, gts012=True, samples=outsamples)
    vcf_query = VCF(vcffile, gts012=True, samples=querysamples)
    # 检查有没有缺少的query个体
    if len(querysamples) > len(vcf_query.samples):
        miss = set(querysamples) - set(vcf_query.samples)
        print(f'query sample miss: {miss}')
        for ind in miss:
            querysamples.remove(ind)
    # 读区间
    regions = []
    if regionfile:
        with open(regionfile) as f:
            for line in f:
                tline = line.strip().split()
                regions.append(f'{tline[0]}:{tline[1]}-{tline[2]}')
    else:
        regions.append(region)
    # 处理vcf
    nregion = 0
    for region in regions:
        print(region)
        df = []
        index = []
        nregion += 1
        if outlist:
            # 定outgroup的allele
            for variant_outgroup, variant_query in zip(vcf_outgroup(region),
                                                       vcf_query(region)):
                if (len(variant_query.ALT) == 1) and (len(variant_query.REF) == 1) and (len(variant_query.ALT[0]) == 1): # 只保留双等位SNP
                    counts = np.bincount(variant_outgroup.gt_types) # 0=HOM_REF, 1=HET, 2=HOM_ALT, 3=UNKNOWN
                    try:
                        major_gt = np.argmax([counts[0], counts[2]]) # 比较0和2哪个多
                    except IndexError: # 没有HOM_ALT
                        major_gt = 0
                    arr = variant_query.gt_types
                    if major_gt == 0:
                        # 如果major allele是alt，则对换ref和alr
                        arr[arr==2] = -9
                        arr[arr==0] = 2
                        arr[arr==-9] = 0
                    df.append(arr.tolist())
                    index.append(variant_query.POS)
        else:
            for variant_query in vcf_query(region):
                if (len(variant_query.ALT) == 1) and (len(variant_query.REF) == 1) and (len(variant_query.ALT[0]) == 1): # 只保留双等位SNP
                    arr = variant_query.gt_types
                    df.append(arr.tolist())
                    index.append(variant_query.POS)

        df = pd.DataFrame(df, columns=vcf_query.samples, index=index)
        df = df.replace(3, np.nan) # 自动 int to float
        # 计算MAF
        freqs = df.sum(axis=1).values / (df.count(axis=1).values * 2)
        freqs[freqs>0.5] = 1-freqs[freqs>0.5]
        df['MAF'] = freqs # 样本ID里要是有叫MAF的那就会被覆盖了！！！
        df['chrom'] = region.split(':')[0]
        df['pos'] = index
        col_order = ['chrom', 'pos', 'MAF'] + querysamples
        df = df[col_order] # 排序
        df = df.loc[df['MAF']>=maf, :] # 按maf筛选
        # 改名
        if rename:
            old2new = loadrename(rename)
            df.columns = [old2new.get(x, x) for x in df.columns]
        # 存excel
        ws = wb.create_sheet(region.replace(':', '_'))
        if not transpose:
            for i, col_name in enumerate(df.columns, 1):
                _ = ws.cell(row=1, column=i, value=col_name)
            n_row, n_col = df.shape
            for row_index in range(n_row):
                for col_index in range(n_col):
                    value = df.iloc[row_index, col_index]
                    cell = ws.cell(row=row_index+2, column=col_index+1, value=value) # excel表是1-index， 第一行是header
                    if col_index > 2: # 前三列是chr,pos,maf
                        if value == 2:
                            color = '7f0000'
                        elif value == 1:
                            color = 'fc8c59'
                        elif value == 0:
                            color = 'fff7ec'
                        else:
                            color = '808080'
                        cell.fill = PatternFill(start_color=color, fill_type = "solid")
            for i in range(n_col):
                if i > 2:
                    ws.column_dimensions[get_column_letter(i+1)].width = 1.25 # 显示一个数字足够了
                elif i == 2:
                    ws.column_dimensions[get_column_letter(i+1)].width = 4.88 # MAF 显示两位小数
                elif i == 0:
                    ws.column_dimensions[get_column_letter(i+1)].width = 2.38 # 染色体列
        if transpose:
            poslist = list(df['pos'])
            del(df['chrom'])
            del(df['MAF'])
            del(df['pos'])
            df = df.T.reset_index() # 转置，header只留一行pos
            df.columns = ['sample'] + poslist #转置并reset_index后，重命名header
            for i, col_name in enumerate(df.columns, 1):
                _ = ws.cell(row=1, column=i, value=col_name)
            n_row, n_col = df.shape
            for row_index in range(n_row):
                for col_index in range(n_col):
                    value = df.iloc[row_index, col_index]
                    cell = ws.cell(row=row_index+2, column=col_index+1, value=value) # excel表是1-index， 第一行是header
                    if col_index > 0: # 前一列是样本ID
                        if value == 2:
                            color = '7f0000'
                        elif value == 1:
                            color = 'fc8c59'
                        elif value == 0:
                            color = 'fff7ec'
                        else:
                            color = '808080'
                        cell.fill = PatternFill(start_color=color, fill_type = "solid")
            for i in range(n_col):
                if i > 0:
                    ws.column_dimensions[get_column_letter(i+1)].width = 1.25 # 显示一个数字足够了
    wb.save(outfile)



if __name__ == '__main__':
    main()
